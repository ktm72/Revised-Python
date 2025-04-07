from openpyxl import load_workbook

workbook = load_workbook(filename='brands_multiple.xlsx')

active_sheet = workbook.active
print(f"active sheet: %s" % active_sheet.title)

for sheet_name in workbook.sheetnames:
  sheet = workbook[sheet_name]
  print(f"sheet name: %s" % sheet_name)

for sheet in workbook.worksheets:
  # sheet = workbook[sheet_name]
  print(f"sheet name: %s" % sheet.title)
  print(sheet.cell(row=2, column=1).value)

workbook.close()