from openpyxl import Workbook

workbook = Workbook()

sheet = workbook.active
sheet.title = "Dummy Table"
sheet2 = workbook.create_sheet("Sheet2")
sheet3 = workbook.create_sheet("Sheet3")
 
# Create headers
headers = ["Column 1", "Column 2", "Column 3", "Column 4"]
sheet.append(headers)
sheet2.append(headers)
sheet3.append(headers)

# Create data
data = [
    ["Data 1-1", "Data 1-2", "Data 1-3", "Data 1-4"],
    ["Data 2-1", "Data 2-2", "Data 2-3", "Data 2-4"],
    ["Data 3-1", "Data 3-2", "Data 3-3", "Data 3-4"],
    ["Data 4-1", "Data 4-2", "Data 4-3", "Data 4-4"],
]

# Append data to the sheet
for row in data:
    sheet.append(row)
    sheet2.append(row)
    sheet3.append(row)


# Save the workbook
workbook.save("example.xlsx")
workbook.close()