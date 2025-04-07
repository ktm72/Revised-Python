from openpyxl import load_workbook

workbook = load_workbook(filename="example.xlsx")
sheet = workbook.active

sheet.cell(row=2, column=2, value="New data");

data_to_write = ["write", "read", "listen", "speak"]

# write data to columns of a row
for column_id, data in enumerate(data_to_write, start=1):
  sheet.cell(row=4, column=column_id, value=data)

# write data to rows of a column
for row_id, data in enumerate(data_to_write, start=2):
  sheet.cell(row=row_id, column=4, value=data)
  
workbook.save('example.xlsx')
workbook.close()