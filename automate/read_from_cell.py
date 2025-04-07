from openpyxl import load_workbook

workbook = load_workbook(filename="brands_output.xlsx")

sheet1 = workbook.active # get the first sheet
print(sheet1.title) # print the title of the sheet

# Iterate through all sheets
for sheet_name in workbook.sheetnames:
	sheet = workbook[sheet_name]
	print(f"Reading sheet: {sheet_name}")

# get the value of the cell A1
cell = sheet1["A1"]
print(f"A1 Cell: {cell.value}")

# get the value of the cell B2
value_in_cell_b2 = sheet1.cell(row=2, column=2).value
print(value_in_cell_b2)

# return None if the cell is empty
# value_in_cell_b10 = sheet1.cell(row=2, column=10).value
# print(f"Empty cell return: {value_in_cell_b10}")


# for row in sheet1.iter_rows(values_only=True):
#     print(row)

for row_num, row in enumerate(sheet1.iter_rows()):
	# print(row)
	for cell_num, cell in enumerate(row):
		print(f"row {row_num}, cell {cell_num} value: {cell.value}")

workbook.close()